# python 3.10.9
# 紀錄各交易所資產(Binance、Bybit、Bitget)
# 作者:0xCatduck
# https://twitter.com/0xCatduck
# 版本: v0.1.0
# 2023/08/29

# 待修正: 多線程查詢；修正查詢方法降低耗時
# 待新增：BingX


import ccxt
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, numbers, Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import time
import os
import os.path

print('----------------------------------')
print('\n\n作者：0xCatduck\nhttps://twitter.com/0xCatduck')
print('\n\n----------------------------------')
# 記錄程式執行的日期時間
now = datetime.now()
dt_string = now.strftime("%Y/%m/%d %H:%M:%S")
print(f"程式執行時間: {dt_string}")
print('----------------------------------')

#######################################################################################

#######################################################################################
# 查詢Bitget的現貨&全倉槓桿&三種合約(U本位、幣本位、USDC)淨值並計算總資產
#######################################################################################
bitget_api_key = os.getenv('bitget_api_key')
bitget_api_secret = os.getenv('bitget_api_secret')
bitget_api_passphrase = os.getenv('bitget_api_passphrase')

# 創建 Bitget 交易所物件
bitget = ccxt.bitget({
    'apiKey': bitget_api_key,
    'secret': bitget_api_secret,

    'password': bitget_api_passphrase,  # 將此處替換為你的 passphrase，沒有則此行移除
    
    'options': {
        'recvWindow': 20000  # 如果跑太慢跳error可以調大這個容許等待值
    }
})


#--------------------------------------------------------------------------------------

# Bitget U本位合約淨值
print('查詢Bitget U本位合約淨值 呱呱喵')
start_time = time.time()

bitget_futures = bitget.private_mix_get_account_accounts({'productType':'UMCBL'})
bitget_futures_usdt_equity = float(bitget_futures['data'][0]['usdtEquity'])

end_time = time.time()

print('Bitget U本位合約查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# Bitget 幣本位合約淨值
print('查詢Bitget 幣本位合約淨值 呱呱喵喵')
start_time = time.time()

bitget_coin = bitget.private_mix_get_account_accounts({'productType':'DMCBL'})

# 初始化淨值之和&幣種列表
bitget_coin_usdt_equity = 0
coin_list = []

# 遍歷帳戶信息，找到 usdtEquity 不為零的幣種及其對應的 usdtEquity 值
for coin in bitget_coin['data']:
    if coin['usdtEquity'] != '0':
        # 將 usdtEquity 的值轉換為浮點數，並加到淨值之和中
        bitget_coin_usdt_equity += float(coin['usdtEquity'])
        # 將幣種及其對應的 usdtEquity 值添加到幣種列表中
        coin_list.append((coin['marginCoin'], coin['usdtEquity']))

end_time = time.time()
print('Bitget 幣本位合約查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# Bitget USDC合約淨值
print('查詢Bitget USDC合約淨值 喵喵喵')
start_time = time.time()

bitget_usdc = bitget.private_mix_get_account_accounts({'productType':'CMCBL'})

# 初始化淨值之和和幣種列表
bitget_usdc_usdt_equity = 0
usdc_list = []

# 遍歷帳戶信息，找到 usdtEquity 不為零的幣種及其對應的 usdtEquity 值
for usdc in bitget_usdc['data']:
    if usdc['usdtEquity'] != '0':
        # 將 usdtEquity 的值轉換為浮點數，並加到淨值之和中
        bitget_usdc_usdt_equity += float(usdc['usdtEquity'])
        # 將幣種及其對應的 usdtEquity 值添加到幣種列表中
        usdc_list.append((usdc['marginCoin'], usdc['usdtEquity']))

end_time = time.time()
print('Bitget USDC合約查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# Bitget 現貨淨值
print('查詢Bitget 現貨淨值 喵喵喵喵')
start_time = time.time()

bitget_spot = bitget.private_spot_get_account_assets()

# 遍歷賬戶信息，找到 available 不為零的幣種及其對應的 available 值
spot_dict = {}
for coin in bitget_spot['data']:
    if coin['available'] != '0.00000000':
        spot_dict[coin['coinName']] = float(coin['available'])

# 獲取所有交易對的最新價格
tickers = bitget.fetch_tickers()

# 遍歷交易對和數量，計算價值
bitget_spot_usdt_equity = 0
for symbol, amount in spot_dict.items():

    if symbol + '/USDT' in tickers:
        # 獲取交易對的最新價格
        ticker = tickers[symbol + '/USDT']
        last_price = ticker['last']
        # 計算價值並加到總價值中
        value = amount * last_price
        bitget_spot_usdt_equity += value

    else:
        # 代表是usdt，直接加入計算
        bitget_spot_usdt_equity += amount

end_time = time.time()
print('Bitget 現貨查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# Bitget 全倉槓桿淨值
print('查詢Bitget 全倉槓桿淨值 喵喵喵喵喵')
start_time = time.time()

bitget_margin = bitget.private_margin_get_cross_account_assets()

# 遍歷賬戶信息，找到 available 不為零的幣種及其對應的 available 值
spot_dict = {}

for coin in bitget_margin['data']:

    if coin['net'] != '0.00000000':

        spot_dict[coin['coin']] = float(coin['net'])

# 獲取所有交易對的最新價格
tickers = bitget.fetch_tickers()

# 遍歷交易對和數量，計算價值
bitget_margin_usdt_equity = 0
for symbol, amount in spot_dict.items():
    if symbol + '/USDT' in tickers:
        # 獲取交易對的最新價格
        ticker = tickers[symbol + '/USDT']
        last_price = ticker['last']
        # 計算價值並加到總價值中
        value = amount * last_price
        bitget_margin_usdt_equity += value
    else:
        # 代表是usdt，直接加入計算
        bitget_margin_usdt_equity += amount

end_time = time.time()
print('Bitget 全倉槓桿查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# 計算 Bitget 總淨值
bitget_assets = bitget_futures_usdt_equity + bitget_coin_usdt_equity + bitget_usdc_usdt_equity + bitget_spot_usdt_equity + bitget_margin_usdt_equity

#######################################################################################

#######################################################################################
# 查詢Bybit資金帳戶、統一交易帳戶、反向合約的淨值並計算總資產
# Bybit的API很多回傳都是用USD計算淨值，所以USDT有折溢價的時候可能看起來會有一點差
# 但林北不管，我全都當USDT，這個鍋Bybit要背
#######################################################################################

bybit_api_key = os.getenv('bybit_api_key')
bybit_api_secret = os.getenv('bybit_api_secret')
# 創建 Bybit 的交易所物件
bybit = ccxt.bybit({
    'apiKey': bybit_api_key,
    'secret': bybit_api_secret,

    'options': {
        'recvWindow': 20000  # 如果跑太慢跳error可以調大這個容許等待值
    }
})


#--------------------------------------------------------------------------------------

# Bybit 統一交易帳戶淨值
print('查詢Bybit 統一交易帳戶淨值 喵喵')
start_time = time.time()

bybit_unified_wallet_balance = bybit.private_get_v5_account_wallet_balance({'accountType':"UNIFIED"})
bybit_unified_usd_equity = float(bybit_unified_wallet_balance['result']['list'][0]['totalEquity']) 

end_time = time.time()
print('Bybit 統一交易帳戶查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# Bybit 反向合約淨值
print('查詢Bybit 反向合約淨值 喵喵喵')
start_time = time.time()

bybit_contract_wallet_balance = bybit.private_get_v5_account_wallet_balance({'accountType':"CONTRACT"})
bybit_contract_wallet_equity = 0
for coin in bybit_contract_wallet_balance['result']['list'][0]['coin']:
    if coin['equity'] != '0':
        ticker = bybit.fetch_ticker(f"{coin['coin']}/USDT")
        coin_price = ticker['last']
        coin_value = float(coin['equity']) * coin_price
        bybit_contract_wallet_equity += coin_value
        # 檢視各資產價值
        #print(f"Coin: {coin['coin']}, Equity: {coin['equity']}, Price: {coin_price}, Value: {coin_value:.2f}")

end_time = time.time()
print('Bybit 反向合約查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# Bybit 資金帳戶淨值
print('查詢Bybit 資金帳戶淨值 呱呱')
start_time = time.time()

bybit_fund_wallet_balance = bybit.private_get_v5_asset_transfer_query_account_coins_balance({'accountType':"FUND"})

bybit_fund_wallet_equity = 0
coin_balances = {}


for coin in bybit_fund_wallet_balance['result']['balance']:
    
    if float(coin['walletBalance']) > 0:
        
        coin_balances[coin['coin']] = coin['walletBalance']
        
        try:
            
            ticker = bybit.fetch_ticker(f"{coin['coin']}USDT")
            coin_price = ticker['last']
            coin_value = float(coin['walletBalance']) * coin_price
            bybit_fund_wallet_equity += coin_value
            
            #檢視各資產價值
            #print(f"{coin['coin']} : {coin['walletBalance']} * {coin_price} = {coin_value:.2f} USDT")
        
        except:
            
            coin_value = float(coin['walletBalance']) * 1
            bybit_fund_wallet_equity += coin_value
            
            #檢視各資產價值
            #print(f"{coin['coin']} : {coin_value:.2f} USDT")

# 檢視持有資產
#print(coin_balances)

end_time = time.time()
print('Bybit 資金帳戶查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# Bybit的金融產品(跟單、交易機器人、理財、餘幣寶等)不支援API查詢，故總資產要額外加入

#--------------------------------------------------------------------------------------

# 計算 Bybit 總淨值
bybit_assets = bybit_unified_usd_equity + bybit_contract_wallet_equity + bybit_fund_wallet_equity



#######################################################################################


#######################################################################################
# 查詢Binance現貨、資金帳戶、理財、U本位合約、幣本位、全倉槓桿的淨值並計算總資產
#######################################################################################

# 創建幣安的交易所物件
binance_api_key = os.getenv('binance_api_key')
binance_api_secret = os.getenv('binance_api_secret')


binance = ccxt.binance({
    'rateLimit': 20,
    'enableRateLimit': True,
    'apiKey': binance_api_key,
    'secret': binance_api_secret,
    'options': {
        'recvWindow': 20000  # 如果跑太慢跳error可以調大這個容許等待值
    }
})

#--------------------------------------------------------------------------------------

# Binance 現貨帳戶淨值
print('查詢Binance 現貨帳戶淨值 呱呱呱呱呱')
start_time = time.time()

balances = binance.fetch_balance()

# 過濾出資產餘額不為 0 的幣種
non_zero_balances = {k: v for k, v in balances['total'].items() if v != 0}

# 這邊包含了活期資產開頭會叫LD(ex.LDBTC)
# 所以過濾出代幣開頭不是 LD (LDO也要拉出來) 的幣種就是現貨資產
binance_spot_balances = {k: v for k, v in non_zero_balances.items() if not k.startswith('LD') or k == 'LDO'}

# 現貨幣種
#print(filtered_balances)

# 定義總價值變數
binance_spot_usdt_equity = 0

# 計算每個代幣的價值並加總
for token, balance in binance_spot_balances.items():
    
    try:
        # 獲取代幣對 USDT 的價格
        symbol = f'{token}/USDT'
        ticker = binance.fetch_ticker(symbol)
        price = ticker['last']

        # 計算代幣價值
        value = balance * price

        # 加總代幣價值
        binance_spot_usdt_equity += value

        # 打印代幣價值
        #print(f'{token} 的價值為 {value} USDT')
    
    # 如果沒有現貨對，則跳過該代幣
    except:
        pass
        #print(f'{token} 沒有現貨對，跳過該代幣')

end_time = time.time()
print('Binance 現貨查詢完成，耗時', end_time - start_time, '秒')


#--------------------------------------------------------------------------------------

# Binance 全倉槓桿淨值
print('查詢Binance 全倉槓桿淨值 呱呱呱呱')
start_time = time.time()

# 獲取帳戶資訊
account_info = binance.sapi_get_margin_account()

# 過濾出 netAsset 不為 0 的幣種
non_zero_assets = [asset for asset in account_info['userAssets'] if float(asset['netAsset']) != 0]
#print(non_zero_assets)


# 定義總價值變數
binance_margin_usdt_equity = 0

# 計算每個幣種的價值並加總
for asset in non_zero_assets:
        try:
            # 獲取幣種對 USDT 的價格
            symbol = f'{asset["asset"]}/USDT'
            ticker = binance.fetch_ticker(symbol)
            price = ticker['last']

            # 計算幣種價值
            value = float(asset['netAsset']) * price
        
        except:
            # 如果沒有查詢到，則代表為USDT，直接加入計算
            value = float(asset['netAsset']) * 1

        # 加總幣種價值
        binance_margin_usdt_equity += value

        # 打印幣種價值
        #print(f'{asset["asset"]} 的價值為 {value} USDT')

end_time = time.time()
print('Binance 全倉槓桿查詢完成，耗時', end_time - start_time, '秒')


#--------------------------------------------------------------------------------------

# Binance U本位合約淨值
print('查詢Binance U本位合約淨值 呱呱呱呱')
start_time = time.time()

# 獲取 totalMarginBalance
binance_futures_usdt_equity = float(binance.fapiprivatev2_get_account()['totalMarginBalance'])

end_time = time.time()
print('Binance U本位合約查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# Binance 幣本位合約淨值
print('查詢Binance 幣本位合約淨值 呱呱呱呱呱呱呱呱呱呱呱呱呱呱呱呱呱')
start_time = time.time()

# 獲取帳戶餘額
balance = binance.dapiprivate_get_balance()

# 過濾出 crossWalletBalance 不為 0 的 asset
filtered_assets = [asset for asset in balance if float(asset['balance']) != 0]

# 定義 binance_coinm_balance 變數
binance_coin_balance = []
binance_coin_usdt_equity = 0
# 計算每個 asset 的價值並加入 binance_coinm_balance
for asset in filtered_assets:
    # 獲取 asset 對 USDT 的價格
    symbol = f'{asset["asset"]}/USDT'
    ticker = binance.fetch_ticker(symbol)
    price = ticker['last']

    # 計算 asset 價值
    value = float(asset['balance']) * price

    # 加入 binance_coinm_balance
    binance_coin_balance.append({'asset': asset['asset'], 'balance': asset['balance'], 'value': value})
    binance_coin_usdt_equity += value

end_time = time.time()
print('Binance 幣本位合約查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# Binance 理財帳戶淨值
# 沒有新幣挖礦或一些產品的查詢API，所以要手動記錄
print('查詢Binance 理財帳戶淨值 旺我是單身狗')
start_time = time.time()

# 活存

# 活存詳細資訊
binance_demand_deposit_detail = binance.sapi_get_lending_union_account()
#print(binance_demand_deposit_detail)
binance_demand_deposit_equity = float(binance_demand_deposit_detail['totalAmountInUSDT'])


# 定存

binance_simple_earn_balance = binance.sapi_get_simple_earn_account()
#print(binance_simple_earn_balance)
binance_time_deposit_equity = float(binance_simple_earn_balance['totalLockedInUSDT'])

binance_earn_equity = binance_time_deposit_equity + binance_demand_deposit_equity

end_time = time.time()
print('Binance 理財查詢完成，耗時', end_time - start_time, '秒')

#--------------------------------------------------------------------------------------

# 計算 Binance 總淨值
binance_assets = binance_spot_usdt_equity + binance_margin_usdt_equity + binance_futures_usdt_equity + binance_coin_usdt_equity + binance_earn_equity


#######################################################################################
print('----------------------------------')
print('查詢完畢')
print('恭喜你！你太幸運啦！你是第一百萬位的使用者(希望真的有很多人用)')
print('現在趕緊的轉錢到開發人員零叉貓鴨的錢包，限時限量打一返三(並沒有)')
print('以下輸出查詢結果並記錄至Excel')
time.sleep(5)
# 輸出所有查詢結果
print('----------------------------------')
print(f"Binance 總資產(只記以下列出者): {binance_spot_usdt_equity + binance_margin_usdt_equity + binance_futures_usdt_equity + binance_coin_usdt_equity + binance_earn_equity:.2f} USDT")
print(f"Binance 現貨餘額: {binance_spot_usdt_equity:.2f} USDT")
print(f"Binance 全倉槓桿餘額: {binance_margin_usdt_equity:.2f} USDT")
print(f"Binance U本位合約餘額: {binance_futures_usdt_equity:.2f} USDT")
print(f"Binance 幣本位合約餘額: {binance_coin_usdt_equity:.2f} USDT")
print(f"Binance 理財餘額: {binance_earn_equity:.2f} USDT")
print('----------------------------------')
# 輸出 Bybit 相關資訊
print(f"Bybit 總資產(只記以下列出者): {bybit_assets:.2f} USDT")
print(f"Bybit 資金帳戶餘額: {bybit_fund_wallet_equity:.2f} USDT")
print(f"Bybit 統一交易帳戶餘額: {bybit_unified_usd_equity:.2f} USDT")
print(f"Bybit 反向合約餘額: {bybit_contract_wallet_equity:.2f} USDT")
print('----------------------------------')
# 輸出 Bitget 相關資訊
print(f"Bitget 總資產(只記以下列出者): {bitget_assets:.2f} USDT")
print(f"Bitget 現貨餘額: {bitget_spot_usdt_equity:.2f} USDT")
print(f"Bitget 全倉槓桿餘額: {bitget_margin_usdt_equity:.2f} USDT")
print(f"Bitget U本位合約餘額: {bitget_futures_usdt_equity:.2f} USDT")
print(f"Bitget 幣本位合約餘額: {bitget_coin_usdt_equity:.2f} USDT")
print(f"Bitget USDC合約餘額: {bitget_usdc_usdt_equity:.2f} USDT")
print('----------------------------------')

#######################################################################################

# 將資料輸入到 Excel 檔案
if os.path.isfile('資產表.xlsx'):
    wb = openpyxl.load_workbook('資產表.xlsx')
    sheet = wb.active  
    row = sheet.max_row + 1
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "CEX Assets"
    row = 1
    sheet['A1'] = '查詢時間'
    sheet['B1'] = '跨交易所\n總資產'
    sheet['C1'] = 'Binance\n總資產'
    sheet['D1'] = 'Binance\n現貨'
    sheet['E1'] = 'Binance\n全倉槓桿'
    sheet['F1'] = 'Binance\nU本位合約'
    sheet['G1'] = 'Binance\n幣本位合約'
    sheet['H1'] = 'Binance\n理財'
    sheet['I1'] = 'Binance\n工人智慧\n手動記錄'
    sheet['J1'] = 'Bybit\n總資產'
    sheet['K1'] = 'Bybit\n資金帳戶'
    sheet['L1'] = 'Bybit\n統一交易\n帳戶'
    sheet['M1'] = 'Bybit\n反向合約'
    sheet['N1'] = 'Bybit\n工人智慧\n手動記錄'
    sheet['O1'] = 'Bitget\n總資產'
    sheet['P1'] = 'Bitget\n現貨'
    sheet['Q1'] = 'Bitget\n全倉槓桿'
    sheet['R1'] = 'Bitget\nU本位合約'
    sheet['S1'] = 'Bitget\n幣本位合約'
    sheet['T1'] = 'Bitget\nUSDC合約'
    sheet['U1'] = 'Bitget\n工人智慧\n手動記錄'
    sheet['V1'] = '備註'

    # 設置標題列的高度
    sheet.row_dimensions[1].height = 60
    # 設置第一列的寬度
    for col in range(1, 2):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 30
    for col in range(2, 22):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 15
    for col in range(22, 23):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 50
    #設置第一列自動換行和置中
    for col in range(1, 23):
        col_letter = chr(col + 64)
        cell = sheet[col_letter + '1']
        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')  
    

    #凍結首行窗格
    sheet.freeze_panes = 'C2'

    row += 1
    wb.save('資產表.xlsx')

# 時間
sheet.cell(row=row, column=1, value=dt_string)
# 總資產
sheet.cell(row=row, column=2, value=(f"=SUM(C{row}+J{row}+O{row})"))
# Binance
sheet.cell(row=row, column=3, value=(f"=SUM(D{row}:I{row})"))
sheet.cell(row=row, column=4, value=round(binance_spot_usdt_equity, 2))
sheet.cell(row=row, column=5, value=round(binance_margin_usdt_equity, 2))
sheet.cell(row=row, column=6, value=round(binance_futures_usdt_equity, 2))
sheet.cell(row=row, column=7, value=round(binance_coin_usdt_equity, 2))
sheet.cell(row=row, column=8, value=round(binance_earn_equity, 2))
# 保留一行給人工記錄
# Bybit
sheet.cell(row=row, column=10, value=(f"=SUM(K{row}:N{row})"))
sheet.cell(row=row, column=11, value=round(bybit_fund_wallet_equity, 2))
sheet.cell(row=row, column=12, value=round(bybit_unified_usd_equity, 2))
sheet.cell(row=row, column=13, value=round(bybit_contract_wallet_equity, 2))
# 保留一行給人工記錄
# Bitget
sheet.cell(row=row, column=15, value=(f"=SUM(P{row}:U{row})"))
sheet.cell(row=row, column=16, value=round(bitget_spot_usdt_equity, 2))
sheet.cell(row=row, column=17, value=round(bitget_margin_usdt_equity, 2))
sheet.cell(row=row, column=18, value=round(bitget_futures_usdt_equity, 2))
sheet.cell(row=row, column=19, value=round(bitget_coin_usdt_equity, 2))
sheet.cell(row=row, column=20, value=round(bitget_usdc_usdt_equity, 2))

# 預設第二列以後的所有行高為40，如果輸出後調整下次不會改到
for row in range(sheet.max_row, sheet.max_row + 1):
    sheet.row_dimensions[row].height = 40

# 設置文字垂直置中
for col in range(1, 23):
    col_letter = chr(col + 64)
    cell = sheet[col_letter + str(row)]
    cell.alignment = Alignment(wrapText=True, vertical='center')  

# 定義整份文件的樣式
default_style = NamedStyle(name='my_style')
default_style.font = Font(name='標楷體', size=14)

# 將樣式應用於整份文件
for sheet in wb.worksheets:
    for row in sheet.rows:
        for cell in row:
            cell.font = default_style.font

#設置顯示小數位
for col in range(2, 22):
    col_letter = chr(col + 64)
    for cell in sheet[col_letter]:
        cell.number_format = numbers.FORMAT_NUMBER_00

# 定義框線樣式
border_style = Border(left=Side(style='dashed', color='000000'),
                      right=Side(style='dashed', color='000000'),
                      top=Side(style='thick', color='000000'),
                      bottom=Side(style='thick', color='000000'))

# 定義重要列樣式
Catduck_style = Border(left=Side(style='thick', color='000000'),
                       right=Side(style='thick', color='000000'),
                       top=Side(style='thick', color='000000'),
                       bottom=Side(style='thick', color='000000'))



# 將框線樣式應用於整份文件
for sheet in wb.worksheets:
    for row in sheet.rows:
        for cell in row:
            cell.border = border_style


### 設置填充顏色及文字顏色 ###

# 填充顏色定義
green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
light_green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
light_blue_fill = PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid')
orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
gray_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
light_gray_fill = PatternFill(start_color='EEECE1', end_color='EEECE1', fill_type='solid')
light_orange_fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')
# 文字顏色定義
red_font = Font(color='FF0000',name='標楷體', size=14)




# 開填

for cell in sheet['A']:
    cell.fill = green_fill
    cell.border = Catduck_style

for cell in sheet['B']:
    cell.font = red_font
    cell.fill = light_gray_fill
    cell.border = Catduck_style

for cell in sheet['C']:
    cell.fill = yellow_fill
    cell.border = Catduck_style

for col in range(4, 10):
    col_letter = chr(col + 64)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill

for cell in sheet['J']:
    cell.fill = orange_fill
    cell.border = Catduck_style

for col in range(11, 15):
    col_letter = chr(col + 64)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill

for cell in sheet['O']:
    cell.fill = light_blue_fill
    cell.border = Catduck_style

for col in range(16, 22):
    col_letter = chr(col + 64)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill

for cell in sheet['V']:
    cell.fill = light_orange_fill
    cell.border = Catduck_style



wb.save('資產表.xlsx')

row_number = row[0].row
print(f"此次更新寫入工作表 {sheet.title} 第 {row_number} 行")