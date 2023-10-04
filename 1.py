# 紀錄各交易所淨值
# Record the net value of each exchange

# python 3.10.9
# v1.0.0 2023/10/04


# 目前支援Binance、Bybit、Bitget、MEXC、BingX、Pionex、Gate.io
# CEXs now support to Binance, Bybit, Bitget, MEXC, BingX, Pionex, Gate.io

# Author: 0xCatduck
# 作者:0xCatduck
# https://twitter.com/0xCatduck
# https://github.com/0xCatduck/assets_between_cex
# https://debank.com/profile/0x77c2113ad732d9e91355838008902cddd4412e56


# 如果這個程式對你有幫助，請順手在github上給我一顆星星，也歡迎請我喝杯咖啡，謝謝！
# If this program is helpful to you, please give me a star on github, feel free to buy me a coffee, thank you!

# 如果有任何問題或發現錯誤，歡迎透過twitter與我聯繫
# If you have any questions or find any bugs, please contact me via twitter

# 使用教學影片連結咐於github上
# Tutorial video link on github

 
#######################################################################################
import ccxt
import threading
import os
from datetime import datetime
import time
import openpyxl
from openpyxl.styles import PatternFill, numbers, Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import requests
import hmac
import hashlib
#######################################################################################

print('--------------------------------------------------------')
print('\n\n作者：0xCatduck\nhttps://twitter.com/0xCatduck')
print('--------------------------------------------------------')
print('恭喜你！你太幸運啦！')
print('你是第一百萬位的使用者(希望真的有很多人用)')
print('現在趕緊的轉錢到開發人員零叉貓鴨的錢包')
print('限時限量打一返三(並沒有)')
print('--------------------------------------------------------\n')
print('順手到下面的連結幫我登入github給個星星')
print('讓我有動力繼續寫些廢物小程式來分享')
print('https://github.com/0xCatduck/assets_between_cex')
print('--------------------------------------------------------\n')

# 記錄程式執行的日期時間
now = datetime.now()
dt_string = now.strftime("%Y/%m/%d %H:%M:%S")
print(f"程式執行時間: {dt_string}")
print('--------------------------------------------------------')


#######################################################################################
# 環境變數設置
#######################################################################################

# 如果沒有以下某間交易所位於環境變數，程式會自動跳過


binance_api_key = os.getenv('binance_api_key')
binance_api_secret = os.getenv('binance_api_secret')

bybit_api_key = os.getenv('bybit_api_key')
bybit_api_secret = os.getenv('bybit_api_secret')

bitget_api_key = os.getenv('bitget_api_key')
bitget_api_secret = os.getenv('bitget_api_secret')
bitget_api_passphrase = os.getenv('bitget_api_passphrase')

mexc_api_key = os.getenv('mexc_api_key')
mexc_api_secret = os.getenv('mexc_api_secret')

bingx_api_key = os.getenv('bingx_api_key')
bingx_api_secret = os.getenv('bingx_api_secret')

pionex_api_key = os.getenv('pionex_api_key')
pionex_api_secret = os.getenv('pionex_api_secret')

gateio_api_key = os.getenv('gateio_api_key')
gateio_api_secret = os.getenv('gateio_api_secret')


#######################################################################################
# 查詢Binance現貨、資金帳戶、理財、U本位合約、幣本位、全倉槓桿、逐倉槓桿的淨值
#######################################################################################


# 創建幣安的交易所物件

if binance_api_key == None or binance_api_secret == None:
    print('未於使用者變數設定Binance api key，跳過紀錄Binance')
    print('--------------------------------------------------------')
    binance = None


else:
    binance = ccxt.binance({
        'rateLimit': 20,
        'enableRateLimit': True,
        'apiKey': binance_api_key,
        'secret': binance_api_secret,
        'options': {
            'recvWindow': 20000  # 如果跑太慢跳error可以調大這個容許等待值
        }
    })

binance_spot_usdt_equity = 0
binance_cross_margin_usdt_equity = 0
binance_query_margin_usdt_equity = 0
binance_futures_usdt_equity = 0
binance_coin_usdt_equity = 0
binance_earn_usdt_equity = 0


# 定義一個函數來計算現貨帳戶淨值
def calculate_binance_spot_equity():
    if binance:
        start_time = time.time()
        global binance_spot_usdt_equity
        balances = binance.fetch_balance()
        non_zero_balances = {k: v for k, v in balances['total'].items() if v != 0}
        binance_spot_balances = {k: v for k, v in non_zero_balances.items() if not k.startswith('LD') or k == 'LDO'}
        for token, balance in binance_spot_balances.items():
            try:
                symbol = f'{token}/USDT'
                ticker = binance.fetch_ticker(symbol)
                price = ticker['last']
                value = balance * price
                binance_spot_usdt_equity += value
            except:
                value = float(balance) * 1
                binance_spot_usdt_equity += value
        end_time = time.time()
        print(f'Binance現貨帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')
# 定義一個函數來計算全倉槓桿淨值
def calculate_binance_cross_margin_equity():
    if binance:
        start_time = time.time()
        global binance_cross_margin_usdt_equity
        account_info = binance.sapi_get_margin_account()
        non_zero_assets = [asset for asset in account_info['userAssets'] if float(asset['netAsset']) != 0]
        for asset in non_zero_assets:
            try:
                symbol = f'{asset["asset"]}/USDT'
                ticker = binance.fetch_ticker(symbol)
                price = ticker['last']
                value = float(asset['netAsset']) * price
            except:
                value = float(asset['netAsset']) * 1
            binance_cross_margin_usdt_equity += value
        end_time = time.time()
        print(f'Binance全倉槓桿淨值查詢時間: {(end_time - start_time):.2f} 秒')

# 定義一個函數來計算逐倉槓桿淨值
def calculate_binance_query_margin_equity():
    global binance_query_margin_usdt_equity  

    if binance:
        start_time = time.time()
        account_info = binance.sapi_get_margin_isolated_account()

        # 取得 totalNetAssetOfBtc 的值
        total_net_asset_of_btc = float(account_info['totalNetAssetOfBtc'])

        # 使用fetch_ticker來獲得BTC對USDT的當前價格
        btc_usdt_ticker = binance.fetch_ticker('BTC/USDT')
        btc_usdt_price = float(btc_usdt_ticker['last'])

        # 計算其在USDT的價值
        binance_query_margin_usdt_equity = total_net_asset_of_btc * btc_usdt_price

        end_time = time.time()
        print(f'Binance 逐倉槓桿帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')

# 定義一個函數來計算 U本位合約淨值
def calculate_binance_futures_equity():
    if binance:
        start_time = time.time()
        global binance_futures_usdt_equity
        # 獲取 Binance U本位合約帳戶資訊
        account_info = binance.fapiprivatev2_get_account()
        binance_futures_usdt_equity += float(account_info['totalMarginBalance'])
        end_time = time.time()
        print(f'Binance U本位合約淨值查詢時間: {(end_time - start_time):.2f} 秒')


# 定義一個函數來計算幣本位合約淨值
def calculate_binance_coin_equity():
    if binance:
        start_time = time.time()
        global binance_coin_usdt_equity
        balance = binance.dapiprivate_get_balance()
        filtered_assets = [asset for asset in balance if float(asset['balance']) != 0]
        for asset in filtered_assets:
            symbol = f'{asset["asset"]}/USDT'
            ticker = binance.fetch_ticker(symbol)
            price = ticker['last']
            value = float(asset['balance']) * price
            binance_coin_usdt_equity += value
        end_time = time.time()
        print(f'Binance 幣本位合約淨值查詢時間: {(end_time - start_time):.2f} 秒')

# 定義一個函數來計算理財帳戶淨值
def calculate_binance_earn_equity():
    if binance:
        start_time = time.time()
        global binance_earn_usdt_equity
        binance_demand_deposit_detail = binance.sapi_get_lending_union_account()
        binance_demand_deposit_equity = float(binance_demand_deposit_detail['totalAmountInUSDT'])
        binance_simple_earn_balance = binance.sapi_get_simple_earn_account()
        binance_time_deposit_equity = float(binance_simple_earn_balance['totalLockedInUSDT'])
        binance_earn_usdt_equity = binance_time_deposit_equity + binance_demand_deposit_equity
        end_time = time.time()
        print(f'Binance 理財帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')


#######################################################################################
# 查詢Bybit資金帳戶、統一交易帳戶、反向合約的淨值
#######################################################################################

# 創建 Bybit 的交易所物件

# 定義全局變量
bybit_unified_usdt_equity = 0
bybit_contract_wallet_usdt_equity = 0
bybit_fund_wallet_usdt_equity = 0

if bybit_api_key == None or bybit_api_secret == None:
    print('未於使用者變數輸入Bybit api key，跳過紀錄Bybit')
    print('--------------------------------------------------------')
    bybit = None

else:
    bybit = ccxt.bybit({
        'apiKey': bybit_api_key,
        'secret': bybit_api_secret,

        'options': {
            'recvWindow': 20000  # 如果跑太慢跳error可以調大這個容許等待值
        }
    })



# 定義一個函數來查詢 Bybit 統一交易帳戶淨值
def caculate_bybit_unified_equity():
    if bybit:
        start_time = time.time()
        global bybit_unified_usdt_equity
        bybit_unified_wallet_balance = bybit.private_get_v5_account_wallet_balance({'accountType':"UNIFIED"})
        bybit_unified_usdt_equity = float(bybit_unified_wallet_balance['result']['list'][0]['totalEquity'])
        end_time = time.time()
        print(f'Bybit 統一交易帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')

# 定義一個函數來查詢 Bybit 反向合約淨值
def caculate_bybit_contract_wallet_usdt_equity():
    if bybit:
        start_time = time.time()
        global bybit_contract_wallet_usdt_equity
        bybit_contract_wallet_balance = bybit.private_get_v5_account_wallet_balance({'accountType':"CONTRACT"})
        for coin in bybit_contract_wallet_balance['result']['list'][0]['coin']:
            if coin['equity'] != '0':
                ticker = bybit.fetch_ticker(f"{coin['coin']}/USDT")
                coin_price = ticker['last']
                coin_value = float(coin['equity']) * coin_price
                bybit_contract_wallet_usdt_equity += coin_value
        end_time = time.time()
        print(f'Bybit 反向合約淨值查詢時間: {(end_time - start_time):.2f} 秒')

# 定義一個函數來查詢 Bybit 資金帳戶淨值
def caculate_bybit_fund_wallet_usdt_balance():
    if bybit:
        start_time = time.time()
        global bybit_fund_wallet_usdt_equity
        bybit_fund_wallet_balance = bybit.private_get_v5_asset_transfer_query_account_coins_balance({'accountType':"FUND"})
        for coin in bybit_fund_wallet_balance['result']['balance']:
            if float(coin['walletBalance']) > 0:
                try:
                    ticker = bybit.fetch_ticker(f"{coin['coin']}USDT")
                    coin_price = ticker['last']
                    coin_value = float(coin['walletBalance']) * coin_price
                    bybit_fund_wallet_usdt_equity += coin_value
                except:
                    coin_value = float(coin['walletBalance']) * 1
                    bybit_fund_wallet_usdt_equity += coin_value
        end_time = time.time()
        print(f'Bybit 資金帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')






#######################################################################################
# 查詢Bitget的現貨&全倉槓桿&三種合約(U本位、幣本位、USDC)淨值
#######################################################################################

# 創建 Bitget 交易所物件

bitget_spot_usdt_equity = 0
bitget_margin_usdt_equity = 0
bitget_futures_usdt_equity = 0
bitget_coin_usdt_equity = 0
bitget_usdc_usdt_equity = 0


if bitget_api_key == None or bitget_api_secret == None or bitget_api_passphrase == None:
    print('未於使用者變數輸入Bitget api key，跳過紀錄Bitget')
    print('--------------------------------------------------------')
    bitget = None

else:

    # 創建 Bitget 交易所物件
    bitget = ccxt.bitget({
        'apiKey': bitget_api_key,
        'secret': bitget_api_secret,

        'password': bitget_api_passphrase,
        
        'options': {
            'recvWindow': 20000  # 如果跑太慢跳error可以調大這個容許等待值
        }
    })




# 定義查詢函數
def caculate_bitget_futures_usdt_equity():
    if bitget:
        start_time = time.time()
        global bitget_futures_usdt_equity
        bitget_futures = bitget.private_mix_get_account_accounts({'productType':'UMCBL'})
        bitget_futures_usdt_equity = float(bitget_futures['data'][0]['usdtEquity'])
        end_time = time.time()
        print(f'Bitget U本位合約淨值查詢時間: {(end_time - start_time):.2f} 秒')

def caculate_bitget_coin_usdt_equity():
    if bitget:
        start_time = time.time()
        global bitget_coin_usdt_equity, coin_list
        bitget_coin = bitget.private_mix_get_account_accounts({'productType':'DMCBL'})
        coin_list = []
        for coin in bitget_coin['data']:
            if coin['usdtEquity'] != '0':
                bitget_coin_usdt_equity += float(coin['usdtEquity'])
                coin_list.append((coin['marginCoin'], coin['usdtEquity']))
        end_time = time.time()
        print(f'Bitget 幣本位合約淨值查詢時間: {(end_time - start_time):.2f} 秒')

def caculate_bitget_usdc_usdt_equity():
    if bitget:
        start_time = time.time()
        global bitget_usdc_usdt_equity, usdc_list
        bitget_usdc = bitget.private_mix_get_account_accounts({'productType':'CMCBL'})
        bitget_usdc_usdt_equity = 0
        usdc_list = []
        for usdc in bitget_usdc['data']:
            if usdc['usdtEquity'] != '0':
                bitget_usdc_usdt_equity += float(usdc['usdtEquity'])
                usdc_list.append((usdc['marginCoin'], usdc['usdtEquity']))
        end_time = time.time()
        print(f'Bitget USDC合約淨值查詢時間: {(end_time - start_time):.2f} 秒')

def caculate_bitget_spot_usdt_equity():
    if bitget:
        start_time = time.time()
        global bitget_spot_usdt_equity
        bitget_spot = bitget.private_spot_get_account_assets()
        spot_dict = {}
        for coin in bitget_spot['data']:
            if coin['available'] != '0.00000000':
                spot_dict[coin['coinName']] = float(coin['available'])
        for symbol, amount in spot_dict.items():
            try:
                symbol_usdt = f'{symbol}/USDT'
                ticker = bitget.fetch_ticker(symbol_usdt)
                last_price = ticker['last']
                value = amount * last_price
                bitget_spot_usdt_equity += value
            except:
                try:
                    symbol_umcbl = f'{symbol}USDT_UMCBL'
                    ticker = bitget.fetch_ticker(symbol_umcbl)
                    last_price = ticker['last']
                    value = amount * last_price
                    bitget_spot_usdt_equity += value
                except:
                    if symbol == 'USDT':
                        bitget_spot_usdt_equity += amount
        end_time = time.time()
        print(f'Bitget 現貨帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')

def caculate_bitget_margin_usdt_equity():
    if bitget:
        start_time = time.time()
        global bitget_margin_usdt_equity
        bitget_margin = bitget.private_margin_get_cross_account_assets()
        margin_dict = {}
        for coin in bitget_margin['data']:
            if coin['net'] != '0.00000000':
                margin_dict[coin['coin']] = float(coin['net'])
        bitget_margin_usdt_equity = 0
        for symbol, amount in margin_dict.items():
            try:
                symbol_usdt = f'{symbol}/USDT'
                ticker = bitget.fetch_ticker(symbol_usdt)
                last_price = ticker['last']
                value = amount * last_price
                bitget_margin_usdt_equity += value
            except:
                try:
                    symbol_umcbl = f'{symbol}USDT_UMCBL'
                    ticker = bitget.fetch_ticker(symbol_umcbl)
                    last_price = ticker['last']
                    value = amount * last_price
                    bitget_margin_usdt_equity += value
                except:
                    if symbol == 'USDT':
                        bitget_margin_usdt_equity += amount

        end_time = time.time()
        print(f'Bitget 全倉槓桿淨值查詢時間:{(end_time - start_time):.2f} 秒')


#######################################################################################
# 查詢MEXC的現貨&合約淨值
#######################################################################################


# 創建 MEXC 交易所物件

mexc_spot_usdt_equity = 0
mexc_futures_usdt_equity = 0


if mexc_api_key == None or mexc_api_secret == None:
    print('未於使用者變數輸入MEXC api key，跳過紀錄MEXC')
    print('--------------------------------------------------------')
    mexc = None

else:

    # 創建 MEXC 交易所物件
    mexc = ccxt.mexc({
        'rateLimit': 20,
        'enableRateLimit': True,
        'apiKey': mexc_api_key,
        'secret': mexc_api_secret,
        'options': {
            'recvWindow': 20000  # 如果跑太慢跳error可以調大這個容許等待值
        }
    })


def caculate_mexc_spot_usdt_equity():
    global mexc_spot_usdt_equity
    result_list = []  # 用於儲存結果

    if mexc:
        start_time = time.time()
        mexc_spot = mexc.fetch_balance()
        total_balances = mexc_spot['total']

        for currency, equity in total_balances.items():
            if equity > 0:  # 僅考慮equity大於0的資產
                if currency == "USDT":
                    usdt_value = equity
                else:
                    ticker = mexc.fetch_ticker(f'{currency}/USDT')
                    latest_price = ticker['last']
                    usdt_value = equity * latest_price

                mexc_spot_usdt_equity += usdt_value  # 累加usdt淨值

                result_list.append({
                    "currency": currency,
                    "currency_equity": equity,
                    "usdt_equity": usdt_value
                })
        
        end_time = time.time()
        print(f'MEXC 現貨帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')

        # 顯示MEXC Spot 個別資產淨值
        # formatted_json = json.dumps(result_list, ensure_ascii=False, indent=4)
        # print(formatted_json)

def caculate_mexc_futures_usdt_equity():
    global mexc_futures_usdt_equity
    result_list = []  # 用於儲存結果

    if mexc:
        start_time = time.time()
        mexc_contract = mexc.contract_private_get_account_assets()
        assets = mexc_contract['data']
        non_zero_equity_assets = [asset for asset in assets if float(asset['equity']) != 0]

        for asset in non_zero_equity_assets:
            currency = asset['currency']
            equity = float(asset['equity'])

            if currency == "USDT":
                usdt_value = equity
            else:
                ticker = mexc.fetch_ticker(f'{currency}/USDT')
                latest_price = ticker['last']
                usdt_value = equity * latest_price

            mexc_futures_usdt_equity += usdt_value  # 累加usdt淨值

            result_list.append({
                "currency": currency,
                "currency_equity": equity,
                "usdt_equity": usdt_value
            })


        end_time = time.time()
        print(f'MEXC 合約帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')

        # 顯示MEXC Futures 個別資產淨值
        # formatted_json = json.dumps(result_list, ensure_ascii=False, indent=4)
        # print(formatted_json)

#######################################################################################
# 查詢BingX的現貨&合約淨值
#######################################################################################

# 創建BingX的交易所物件
bingx_spot_usdt_equity = 0
bingx_futures_usdt_equity = 0

if bingx_api_key == None or bingx_api_secret == None:
    print('未於使用者變數輸入BingX api key，跳過紀錄BingX')
    print('--------------------------------------------------------')
    bingx = None

else:
    bingx = ccxt.bingx({
        'rateLimit': 20,
        'enableRateLimit': True,
        'apiKey': bingx_api_key,
        'secret': bingx_api_secret,
        'options': {
            'recvWindow': 20000  # 如果跑太慢跳error可以調大這個容許等待值
        }
    })

def caculate_bingx_spot_usdt_equity():
    start_time = time.time()
    global bingx_spot_usdt_equity
    balance = bingx.fetch_balance()
    totals = balance['total']

    # 排除 'VST' 資產
    if 'VST' in totals:
        del totals['VST']

    # 初始化 USDT equity
    bingx_spot_usdt_equity = totals['USDT']

    # 對於每個非USDT的資產，獲取其對USDT的價格並計算其USDT等值
    for asset, amount in totals.items():
        if asset != 'USDT':
            ticker = bingx.fetch_ticker(f'{asset}/USDT:USDT')
            price = ticker['last']  # 使用最後的交易價格
            bingx_spot_usdt_equity += amount * price
    end_time = time.time()
    print(f'BingX 現貨帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')

def caculate_bingx_futures_usdt_equity():
    start_time = time.time()
    global bingx_futures_usdt_equity
    response = bingx.swap_v2_private_get_user_balance()
    bingx_futures_usdt_equity = float(response['data']['balance']['equity'])
    end_time = time.time()
    print(f'BingX 合約帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')
    # 只含永續合約，不含標準合約

#######################################################################################
# 查詢派網Pionex的主帳戶&合約淨值
#######################################################################################

# 創建 Pionex 交易所物件
pionex_spot_usdt_equity = 0
pionex_futures_usdt_equity = 0

if pionex_api_key == None or pionex_api_secret == None:
    print('未於使用者變數輸入Pionex api key，跳過紀錄Pionex')
    print('--------------------------------------------------------')
    pionex = None

else:
    pionex = True


def fetch_from_pionex(path_url, params=None):
    base_url = 'https://api.pionex.com'
    
    pionex_api_key = os.getenv('pionex_api_key')
    pionex_api_secret = os.getenv('pionex_api_secret')
    
    timestamp = str(int(time.time() * 1000))  # 毫秒時間戳
    query = f'timestamp={timestamp}'

    if params:
        for key, value in params.items():
            query += f'&{key}={value}'
    
    url = f'{base_url}{path_url}?{query}'
    
    headers = {
        'Content-Type': 'application/json',
        'PIONEX-KEY': pionex_api_key
    }

    method = 'GET'
    sorted_params = '&'.join(sorted(query.split('&')))
    path_url_with_query = path_url + '?' + sorted_params

    signature = hmac.new(
        pionex_api_secret.encode(),
        (method + path_url_with_query).encode(),
        hashlib.sha256
    ).hexdigest()

    headers['PIONEX-SIGNATURE'] = signature

    try:
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            return response.json()
        else:
            print(f'錯誤：無法取得資料，狀態碼：{response.status_code}')
            return None
    except Exception as e:
        print(f'錯誤：{str(e)}')
        return None
        
def fetch_pionex_trades(symbol):
    path_url = '/api/v1/market/trades'
    params = {
        'symbol': f'{symbol.upper()}_USDT',
        'limit': 1
    }
    trade_info = fetch_from_pionex(path_url, params)

    if trade_info and 'data' in trade_info and 'trades' in trade_info['data'] and trade_info['data']['trades']:
        return float(trade_info['data']['trades'][0]['price'])
    else:
        return None




def calculate_pionex_spot_usdt_equity():
    start_time = time.time()
    global pionex_spot_usdt_equity
    path_url = '/api/v1/account/balances'
    balance_info = fetch_from_pionex(path_url)
    pionex_spot_usdt_equity = 0.0

    if balance_info and 'result' in balance_info and balance_info['result'] and 'data' in balance_info and 'balances' in balance_info['data']:
        for item in balance_info['data']['balances']:
            coin = item['coin']
            amount = float(item['free'])  # 這裡只計算free的量，如果需要包括frozen，請進行調整

            if coin == 'USDT':
                pionex_spot_usdt_equity += amount
            else:
                price = fetch_pionex_trades(coin)
                if price:
                    added_equity = amount * price 
                    pionex_spot_usdt_equity += added_equity
    end_time = time.time()
    print(f'Pionex 現貨帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')

def calculate_pionex_futures_usdt_equity():
    start_time = time.time()
    global pionex_futures_usdt_equity

    path_url = '/uapi/v1/account/detail'
    balance_info = fetch_from_pionex(path_url)

    pionex_futures_usdt_equity = 0.0

    if balance_info and 'result' in balance_info and balance_info['result'] and 'data' in balance_info and 'balances' in balance_info['data']:
        for item in balance_info['data']['balances']:
            assets = float(item['assets'])

            # 只有當 assets 不為 0 時，才進行計算
            if assets != 0:
                coin = item['coin']
                free = float(item['free'])
                unrealizedPnL = float(item['unrealizedPnL'])
                balance = free + unrealizedPnL

                if coin == 'USDT':
                    pionex_futures_usdt_equity += balance
                else:
                    price = fetch_pionex_trades(coin)
                    if price:
                        pionex_futures_usdt_equity += balance * price
    end_time = time.time()
    print(f'Pionex 合約帳戶淨值查詢時間: {(end_time - start_time):.2f} 秒')


#######################################################################################
# 查詢Gate.io的主帳戶&合約淨值
#######################################################################################

# 創建 Gate.io 交易所物件
gateio_cbbc_usdt_equity = 0 # 輪證帳戶
gateio_delivery_usdt_equity = 0 # 交割合約
gateio_finance_usdt_equity = 0 # 理財帳戶
gateio_futures_usdt_equity = 0 # 永續合約
gateio_margin_usdt_equity = 0 # 逐倉槓桿帳戶
gateio_options_usdt_equity = 0 # 期權帳戶
gateio_payment_usdt_equity = 0 # 支付帳戶
gateio_quant_usdt_equity = 0 # 跟單帳戶
gateio_spot_usdt_equity = 0 # 現貨帳戶
gateio_total_usdt_equity = 0 # 總淨值

if gateio_api_key == None or gateio_api_secret == None:
    print('未於使用者變數輸入Gate.io api key，跳過紀錄Gate.io')
    print('--------------------------------------------------------')
    gateio = None

else:
    gateio = ccxt.gateio({
        'rateLimit': 20,
        'enableRateLimit': True,
        'apiKey': gateio_api_key,
        'secret': gateio_api_secret,
        'options': {
            'recvWindow': 20000  # 如果跑太慢跳error可以調大這個容許等待值
        }
    })

def calculate_gateio_usdt_equity():
    start_time = time.time()
    global gateio_cbbc_usdt_equity
    global gateio_delivery_usdt_equity
    global gateio_finance_usdt_equity
    global gateio_futures_usdt_equity
    global gateio_margin_usdt_equity
    global gateio_options_usdt_equity
    global gateio_payment_usdt_equity
    global gateio_quant_usdt_equity
    global gateio_spot_usdt_equity
    global gateio_total_usdt_equity


    # 假設gateio.private_wallet_get_total_balance()是這樣的結果
    result = gateio.private_wallet_get_total_balance()
    
    if 'details' in result:
        details = result['details']
        
        gateio_cbbc_usdt_equity = float(details['cbbc']['amount'])
        gateio_delivery_usdt_equity = float(details['delivery']['amount'])
        gateio_finance_usdt_equity = float(details['finance']['amount'])
        gateio_futures_usdt_equity = float(details['futures']['amount'])
        gateio_margin_usdt_equity = float(details['margin']['amount'])
        gateio_options_usdt_equity = float(details['options']['amount'])
        gateio_payment_usdt_equity = float(details['payment']['amount'])
        gateio_quant_usdt_equity = float(details['quant']['amount'])
        gateio_spot_usdt_equity = float(details['spot']['amount'])

    if 'total' in result:
        gateio_total_usdt_equity = float(result['total']['amount'])

    end_time = time.time()
    print(f'Gate.io 全帳戶查詢時間: {(end_time - start_time):.2f} 秒')


#######################################################################################

# 多線程同時執行不同計算函數


exchanges_data = {
    'binance': {
        'functions': [
            calculate_binance_spot_equity,
            calculate_binance_cross_margin_equity,
            calculate_binance_query_margin_equity,
            calculate_binance_futures_equity,
            calculate_binance_coin_equity,
            calculate_binance_earn_equity
        ],
        'assets': [
            binance_spot_usdt_equity,
            binance_cross_margin_usdt_equity,
            binance_query_margin_usdt_equity,
            binance_futures_usdt_equity,
            binance_coin_usdt_equity,
            binance_earn_usdt_equity
        ],
        'labels': [
            'Binance spot balance:', 'Binance cross margin balance:',
            'Binance query margin balance:', 'Binance U-based contract balance:',
            'Binance coin-based contract balance:', 'Binance earn balance:'
        ]
    },
    'bybit': {
        'functions': [
            caculate_bybit_unified_equity,
            caculate_bybit_contract_wallet_usdt_equity,
            caculate_bybit_fund_wallet_usdt_balance
        ],
        'assets': [
            bybit_unified_usdt_equity,
            bybit_contract_wallet_usdt_equity,
            bybit_fund_wallet_usdt_equity
        ],
        'labels': [
            'Bybit fund account balance:', 'Bybit unified account balance:',
            'Bybit inverse contract balance:'
        ]
    },
    'bitget': {
        'functions': [
            caculate_bitget_futures_usdt_equity,
            caculate_bitget_coin_usdt_equity,
            caculate_bitget_usdc_usdt_equity,
            caculate_bitget_spot_usdt_equity,
            caculate_bitget_margin_usdt_equity
        ],
        'assets': [
            bitget_futures_usdt_equity,
            bitget_coin_usdt_equity,
            bitget_usdc_usdt_equity,
            bitget_spot_usdt_equity,
            bitget_margin_usdt_equity
        ],
        'labels': [
            'Bitget futures account balance:', 'Bitget coin account balance:',
            'Bitget USDC contract balance:', 'Bitget spot balance:',
            'Bitget margin balance:'
        ]
    },
    'mexc': {
        'functions': [
            caculate_mexc_spot_usdt_equity,
            caculate_mexc_futures_usdt_equity
        ],
        'assets': [
            mexc_spot_usdt_equity,
            mexc_futures_usdt_equity
        ],
        'labels': [
            'MEXC spot balance:', 'MEXC futures balance:'
        ]
    },
    'bingx': {
        'functions': [
            caculate_bingx_spot_usdt_equity,
            caculate_bingx_futures_usdt_equity
        ],
        'assets': [
            bingx_spot_usdt_equity,
            bingx_futures_usdt_equity
        ],
        'labels': [
            'BingX spot balance:', 'BingX futures balance:'
        ]
    },
    'pionex': {
        'functions': [
            calculate_pionex_spot_usdt_equity,
            calculate_pionex_futures_usdt_equity
        ],
        'assets': [
            pionex_spot_usdt_equity,
            pionex_futures_usdt_equity
        ],
        'labels': [
            'Pionex spot balance:', 'Pionex futures balance:'
        ]
    },
    'gateio': {
        'functions': [
            calculate_gateio_usdt_equity
        ],
        'assets': [
            gateio_cbbc_usdt_equity,
            gateio_delivery_usdt_equity,
            gateio_finance_usdt_equity,
            gateio_futures_usdt_equity,
            gateio_margin_usdt_equity,
            gateio_options_usdt_equity,
            gateio_payment_usdt_equity,
            gateio_quant_usdt_equity,
            gateio_spot_usdt_equity,
            gateio_total_usdt_equity
        ],
        'labels': [
            'Gate.io cbbc balance:', 'Gate.io delivery balance:',
            'Gate.io finance balance:', 'Gate.io futures balance:',
            'Gate.io margin balance:', 'Gate.io options balance:',
            'Gate.io payment balance:', 'Gate.io quant balance:',
            'Gate.io spot balance:', 'Gate.io total balance:'
        ],
        'has_total': True
    }
}

threads = {}

# 初始化線程
for exchange, data in exchanges_data.items():
    if globals().get(exchange):
        threads[exchange] = [threading.Thread(target=func) for func in data["functions"]]

# 啟動和等待所有線程
for t_list in threads.values():
    for t in t_list:
        t.start()

for t_list in threads.values():
    for t in t_list:
        t.join()





#######################################################################################
# 定義交易所的資料

exchanges_info = {
    'binance': {
        'assets': [
            binance_spot_usdt_equity, binance_cross_margin_usdt_equity,
            binance_query_margin_usdt_equity, binance_futures_usdt_equity,
            binance_coin_usdt_equity, binance_earn_usdt_equity
        ],
        'labels': [
            'Binance spot balance:', 'Binance cross margin balance:',
            'Binance query margin balance:', 'Binance U-based contract balance:',
            'Binance coin-based contract balance:', 'Binance earn balance:'
        ]
    },
    'bybit': {
        'assets': [
            bybit_unified_usdt_equity, bybit_contract_wallet_usdt_equity, 
            bybit_fund_wallet_usdt_equity
        ],
        'labels': [
            'Bybit fund account balance:', 'Bybit unified account balance:',
            'Bybit inverse contract balance:'
        ]
    },
    'bitget': {
        'assets': [
            bitget_futures_usdt_equity, bitget_coin_usdt_equity,
            bitget_usdc_usdt_equity, bitget_spot_usdt_equity,
            bitget_margin_usdt_equity
        ],
        'labels': [
            'Bitget futures account balance:', 'Bitget coin account balance:',
            'Bitget USDC contract balance:', 'Bitget spot balance:',
            'Bitget margin balance:'
        ]
    },
    'mexc': {
        'assets': [
            mexc_spot_usdt_equity, mexc_futures_usdt_equity
        ],
        'labels': [
            'MEXC spot balance:', 'MEXC futures account balance:'
        ]
    },
    'bingx': {
        'assets': [
            bingx_spot_usdt_equity, bingx_futures_usdt_equity
        ],
        'labels': [
            'BingX spot balance:', 'BingX futures account balance:'
        ]
    },
    'pionex': {
        'assets': [
            pionex_spot_usdt_equity, pionex_futures_usdt_equity
        ],
        'labels': [
            'Pionex spot balance:', 'Pionex futures account balance:'
        ]
    },
    'gateio': {
        'assets': [
            gateio_spot_usdt_equity, gateio_margin_usdt_equity, gateio_futures_usdt_equity, 
            gateio_delivery_usdt_equity, gateio_quant_usdt_equity, gateio_cbbc_usdt_equity, 
            gateio_finance_usdt_equity, gateio_options_usdt_equity, gateio_payment_usdt_equity
        ],
        'labels': [
            'Gate.io spot balance:', 'Gate.io query margin balance:', 'Gate.io futures balance:', 
            'Gate.io delivery balance:', 'Gate.io quant balance:', 'Gate.io cbbc balance:',
            'Gate.io finance balance:', 'Gate.io options balance:', 'Gate.io payment balance:'
        ],
        'has_total': True
    }
}

# 計算總淨值
total_assets = {}
for exchange, info in exchanges_info.items():
    exchange_obj = globals()[exchange] if exchange in globals() else None
    if info.get('has_total'):
        total_assets[exchange] = gateio_total_usdt_equity  
    else:
        is_queried = exchange_obj is not None
        total_assets[exchange] = sum(info['assets']) if is_queried else 0

cross_cexs_assets = sum(total_assets.values())

print('\n--------------------------------------------------------')
print('                 以下輸出此次查詢結果')
print('--------------------------------------------------------\n')
print('--------------------------------------------------------')
print(f'{"Cross CEXs total assets:":<40} {cross_cexs_assets:>10.2f} USDT')
print('--------------------------------------------------------\n')

# 輸出每一個交易所的資訊
for exchange, info in exchanges_info.items():
    print(f'{exchange.capitalize()} total assets:'.ljust(40), f'{total_assets[exchange]:>10.2f} USDT')
    for asset, label in zip(info['assets'], info['labels']):
        print(label.ljust(40), f'{asset:>10.2f} USDT')
    print('--------------------------------------------------------\n')




#######################################################################################

# 將資料輸入到 Excel 檔案
if os.path.isfile('crypto_assets.xlsx'):
    wb = openpyxl.load_workbook('crypto_assets.xlsx')
    sheet = wb.active  
    row = sheet.max_row + 1
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "CEX Assets"
    row = 1
    sheet.cell(row=1, column=1).value = '查詢時間' # A1
    sheet.cell(row=1, column=2).value = '跨交易所\n總資產' # B1
    sheet.cell(row=1, column=3).value = 'Binance\n總資產' # C1
    sheet.cell(row=1, column=4).value = 'Binance\n現貨' # D1
    sheet.cell(row=1, column=5).value = 'Binance\n全倉槓桿' # E1
    sheet.cell(row=1, column=6).value = 'Binance\n逐倉槓桿' # F1
    sheet.cell(row=1, column=7).value = 'Binance\nU本位合約' # G1
    sheet.cell(row=1, column=8).value = 'Binance\n幣本位合約' # H1
    sheet.cell(row=1, column=9).value = 'Binance\n理財' # I1
    sheet.cell(row=1, column=10).value = 'Binance\n工人智慧\n手動記錄' # J1
    sheet.cell(row=1, column=11).value = 'Binance\n工人智慧\n備註' # K1
    sheet.cell(row=1, column=12).value = 'Bybit\n總資產' # L1
    sheet.cell(row=1, column=13).value = 'Bybit\n資金帳戶' # M1
    sheet.cell(row=1, column=14).value = 'Bybit\n統一交易\n帳戶' # N1
    sheet.cell(row=1, column=15).value = 'Bybit\n反向合約' # O1
    sheet.cell(row=1, column=16).value = 'Bybit\n工人智慧\n手動記錄' # P1
    sheet.cell(row=1, column=17).value = 'Bybit\n工人智慧\n備註' # Q1
    sheet.cell(row=1, column=18).value = 'Bitget\n總資產' # R1
    sheet.cell(row=1, column=19).value = 'Bitget\n現貨' # S1
    sheet.cell(row=1, column=20).value = 'Bitget\n全倉槓桿' # T1
    sheet.cell(row=1, column=21).value = 'Bitget\nU本位合約' # U1
    sheet.cell(row=1, column=22).value = 'Bitget\n幣本位合約' # V1
    sheet.cell(row=1, column=23).value = 'Bitget\nUSDC合約' # W1
    sheet.cell(row=1, column=24).value = 'Bitget\n工人智慧\n手動記錄' # X1
    sheet.cell(row=1, column=25).value = 'Bitget\n工人智慧\n備註' # Y1
    sheet.cell(row=1, column=26).value = 'MEXC\n總資產' # Z1
    sheet.cell(row=1, column=27).value = 'MEXC\n現貨' # AA1
    sheet.cell(row=1, column=28).value = 'MEXC\n合約' # AB1
    sheet.cell(row=1, column=29).value = 'MEXC\n工人智慧\n手動記錄' # AC1
    sheet.cell(row=1, column=30).value = 'MEXC\n工人智慧\n備註' # AD1
    sheet.cell(row=1, column=31).value = 'BingX\n總資產' # AE1
    sheet.cell(row=1, column=32).value = 'BingX\n現貨' # AF1
    sheet.cell(row=1, column=33).value = 'BingX\n合約' # AG1
    sheet.cell(row=1, column=34).value = 'BingX\n工人智慧\n手動記錄' # AH1
    sheet.cell(row=1, column=35).value = 'BingX\n工人智慧\n備註' # AI1
    sheet.cell(row=1, column=36).value = 'Pionex\n總資產' # AJ1
    sheet.cell(row=1, column=37).value = 'Pionex\n現貨' # AK1
    sheet.cell(row=1, column=38).value = 'Pionex\n合約' # AL1
    sheet.cell(row=1, column=39).value = 'Pionex\n工人智慧\n手動記錄' # AM1
    sheet.cell(row=1, column=40).value = 'Pionex\n工人智慧\n備註' # AN1
    sheet.cell(row=1, column=41).value = 'Gate.io\n總資產' # AO1
    sheet.cell(row=1, column=42).value = 'Gate.io\n現貨' # AP1
    sheet.cell(row=1, column=43).value = 'Gate.io\n逐倉槓桿' # AQ1
    sheet.cell(row=1, column=44).value = 'Gate.io\nU本位合約' # AR1
    sheet.cell(row=1, column=45).value = 'Gate.io\n交割合約' # AS1
    sheet.cell(row=1, column=46).value = 'Gate.io\n跟單帳戶' # AT1
    sheet.cell(row=1, column=47).value = 'Gate.io\n輪證帳戶' # AU1
    sheet.cell(row=1, column=48).value = 'Gate.io\n理財帳戶' # AV1
    sheet.cell(row=1, column=49).value = 'Gate.io\n期權帳戶' # AW1
    sheet.cell(row=1, column=50).value = 'Gate.io\n支付帳戶' # AX1
    sheet.cell(row=1, column=51).value = 'Gate.io\n工人智慧\n手動記錄' # AY1
    sheet.cell(row=1, column=52).value = 'Gate.io\n工人智慧\n備註' # AZ1
    sheet.cell(row=1, column=53).value = '閒雜人等\n資產紀錄1' # BA1
    sheet.cell(row=1, column=54).value = '閒雜人等\n資產紀錄1\n備註' # BB1
    sheet.cell(row=1, column=55).value = '閒雜人等\n資產紀錄2' # BC1
    sheet.cell(row=1, column=56).value = '閒雜人等\n資產紀錄2\n備註' # BD1
    sheet.cell(row=1, column=57).value = '閒雜人等\n資產紀錄3' # BE1
    sheet.cell(row=1, column=58).value = '閒雜人等\n資產紀錄3\n備註' # BF1
    sheet.cell(row=1, column=59).value = '總備註' # BG1






    # 設置標題列的高度
    sheet.row_dimensions[1].height = 60
    # 設置第一列的寬度
    for col in range(1, 2):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 30
    for col in range(2, 60):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 16
    for col in range(1, sheet.max_column + 1):  # 從第一列到最後一列
        cell_value = sheet.cell(row=1, column=col).value
        if "備註" in str(cell_value):  # 如果該列的值包含"備註"
            col_letter = get_column_letter(col)
            sheet.column_dimensions[col_letter].width = 50
    #設置第一列自動換行和置中
    for col in range(1, 60):
        col_letter = get_column_letter(col)
        cell = sheet[col_letter + '1']
        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')  


    #凍結首行窗格
    sheet.freeze_panes = 'C2'

    row += 1
    wb.save('crypto_assets.xlsx')






# 時間
sheet.cell(row=row, column=1, value=dt_string)


# 總資產
total_assets = f"=C{row}+L{row}+R{row}+Z{row}+AE{row}+AJ{row}+AO{row}+BA{row}+BC{row}+BE{row}"
sheet.cell(row=row, column=2, value=str(total_assets))

# Binance
binance_sum_formula = f"=SUM(D{row}:J{row})"
sheet.cell(row=row, column=3).value = binance_sum_formula
sheet.cell(row=row, column=4, value=round(binance_spot_usdt_equity, 2))
sheet.cell(row=row, column=5, value=round(binance_cross_margin_usdt_equity, 2))
sheet.cell(row=row, column=6, value=round(binance_query_margin_usdt_equity, 2))
sheet.cell(row=row, column=7, value=round(binance_futures_usdt_equity, 2))
sheet.cell(row=row, column=8, value=round(binance_coin_usdt_equity, 2))
sheet.cell(row=row, column=9, value=round(binance_earn_usdt_equity, 2))
sheet.cell(row=row, column=10,value=None)
sheet.cell(row=row, column=11,value=None)


# Bybit
bybit_sum_formula = f"=SUM(M{row}:P{row})"
sheet.cell(row=row, column=12, value=bybit_sum_formula)
sheet.cell(row=row, column=13, value=round(bybit_fund_wallet_usdt_equity, 2))
sheet.cell(row=row, column=14, value=round(bybit_unified_usdt_equity, 2))
sheet.cell(row=row, column=15, value=round(bybit_contract_wallet_usdt_equity, 2))
sheet.cell(row=row, column=16,value=None)
sheet.cell(row=row, column=17,value=None)


# Bitget
bitget_sum_formula = f"=SUM(S{row}:X{row})"
sheet.cell(row=row, column=18, value=bitget_sum_formula)
sheet.cell(row=row, column=19, value=round(bitget_spot_usdt_equity, 2))
sheet.cell(row=row, column=20, value=round(bitget_margin_usdt_equity, 2))
sheet.cell(row=row, column=21, value=round(bitget_futures_usdt_equity, 2))
sheet.cell(row=row, column=22, value=round(bitget_coin_usdt_equity, 2))
sheet.cell(row=row, column=23, value=round(bitget_usdc_usdt_equity, 2))
sheet.cell(row=row, column=24,value=None)
sheet.cell(row=row, column=25,value=None)

# MEXC
mexc_sum_formula = f"=SUM(AA{row}:AC{row})"
sheet.cell(row=row, column=26, value=mexc_sum_formula)
sheet.cell(row=row, column=27, value=round(mexc_spot_usdt_equity, 2))
sheet.cell(row=row, column=28, value=round(mexc_futures_usdt_equity, 2))
sheet.cell(row=row, column=29,value=None)
sheet.cell(row=row, column=30,value=None)

# BingX
bingx_sum_formula = f"=SUM(AF{row}:AH{row})"
sheet.cell(row=row, column=31, value=bingx_sum_formula)
sheet.cell(row=row, column=32, value=round(bingx_spot_usdt_equity, 2))
sheet.cell(row=row, column=33, value=round(bingx_futures_usdt_equity, 2))
sheet.cell(row=row, column=34,value=None)
sheet.cell(row=row, column=35,value=None)

# Pionex
pionex_sum_formula = f"=SUM(AK{row}:AM{row})"
sheet.cell(row=row, column=36, value=pionex_sum_formula)
sheet.cell(row=row, column=37, value=round(pionex_spot_usdt_equity, 2))
sheet.cell(row=row, column=38, value=round(pionex_futures_usdt_equity, 2))
sheet.cell(row=row, column=39,value=None)
sheet.cell(row=row, column=40,value=None)

# Gate.io
gateio_sum_formula = f"=SUM(AP{row}:AY{row})"
sheet.cell(row=row, column=41, value=gateio_sum_formula)
sheet.cell(row=row, column=42, value=round(gateio_spot_usdt_equity, 2))
sheet.cell(row=row, column=43, value=round(gateio_margin_usdt_equity, 2))
sheet.cell(row=row, column=44, value=round(gateio_futures_usdt_equity, 2))
sheet.cell(row=row, column=45, value=round(gateio_delivery_usdt_equity, 2))
sheet.cell(row=row, column=46, value=round(gateio_quant_usdt_equity, 2))
sheet.cell(row=row, column=47, value=round(gateio_cbbc_usdt_equity, 2))
sheet.cell(row=row, column=48, value=round(gateio_finance_usdt_equity, 2))
sheet.cell(row=row, column=49, value=round(gateio_options_usdt_equity, 2))
sheet.cell(row=row, column=50, value=round(gateio_payment_usdt_equity, 2))
sheet.cell(row=row, column=51,value=None)
sheet.cell(row=row, column=52,value=None)






###################### 樣式設定 ######################


# 預設第二列以後的所有行高為35，如果輸出後調整下次不會改到
for row in range(sheet.max_row, sheet.max_row + 1):
    sheet.row_dimensions[row].height = 35

# 設置文字垂直置中
for col in range(1, 60):
    col_letter = get_column_letter(col)
    cell = sheet[col_letter + str(row)]
    cell.alignment = Alignment(wrapText=True, vertical='center')  

# 定義整份文件的樣式
default_style = NamedStyle(name='my_style')
default_style.font = Font(name='標楷體', size=14)

# 將樣式應用於整份文件
for sheet in wb.worksheets:
    for rows in sheet.rows:
        for cell in rows:
            cell.font = default_style.font

#設置顯示小數位
for col in range(2, 60):
    col_letter = get_column_letter(col)
    for cell in sheet[col_letter]:
        cell.number_format = numbers.FORMAT_NUMBER_00

# 定義框線樣式
border_style = Border(left=Side(style='dashed', color='000000'),
                      right=Side(style='dashed', color='000000'),
                      top=Side(style='thick', color='000000'),
                      bottom=Side(style='thick', color='000000'))

# 定義重要列樣式
Catduck_style = Border(left=Side(style='thick', color='000000'),
                       right=Side(style='thick', color='000000'),
                       top=Side(style='thick', color='000000'),
                       bottom=Side(style='thick', color='000000'))



# 將框線樣式應用於整份文件
for sheet in wb.worksheets:
    for rows in sheet.rows:
        for cell in rows:
            cell.border = border_style



### 設置填充顏色及文字顏色 ###

# 填充顏色定義
green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid') # 綠色
light_green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid') # 淺綠色
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # 黃色
light_blue_fill = PatternFill(start_color='BFEFFF', end_color='BFEFFF', fill_type='solid') # 淺藍色
orange_fill = PatternFill(start_color='F7A600', end_color='F7A600', fill_type='solid') # Bybit橘色
gray_fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid') # 灰色
light_gray_fill = PatternFill(start_color='EEECE1', end_color='EEECE1', fill_type='solid') # 淺灰色
light_orange_fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid') # 淺橘色
blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid') # 藍色
light_purple_fill = PatternFill(start_color='E6B9EC', end_color='E6B9EC', fill_type='solid') # 淺紫色
water_blue_fill = PatternFill(start_color='33CCFF', end_color='33CCFF', fill_type='solid') # 水藍色
bright_orange_fill = PatternFill(start_color='FF7028', end_color='FF7028', fill_type='solid') # Pionex橘色
lime_green_fill= PatternFill(start_color='66EC83', end_color='66EC83', fill_type='solid') # 淺綠色

# 文字顏色定義
red_font = Font(color='FF0000',name='標楷體', size=14) # 紅色



# 時間
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=1) # 時間, A列
    cell.fill = green_fill
    cell.border = Catduck_style

# 總資產
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=2) # 總資產, B列
    cell.font = red_font
    cell.fill = light_gray_fill
    cell.border = Catduck_style

# Binance
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=3) # Binance 總資產, C列
    cell.fill = yellow_fill
    cell.border = Catduck_style
for col in range(4, 12): 
    col_letter = get_column_letter(col)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill

# Bybit
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=12) # Bybit 總資產, L列
    cell.fill = orange_fill
    cell.border = Catduck_style
for col in range(13, 18):
    col_letter = get_column_letter(col)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill

# Bitget
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=18) # Bitget 總資產, R列
    cell.fill = water_blue_fill
    cell.border = Catduck_style
for col in range(19, 26):
    col_letter = get_column_letter(col)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill

# MEXC
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=26) # MEXC 總資產, Z列
    cell.fill = blue_fill
    cell.border = Catduck_style
for col in range(27, 31):
    col_letter = get_column_letter(col)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill
    
# BingX
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=31) # BingX 總資產, AE列
    cell.fill = light_blue_fill
    cell.border = Catduck_style
for col in range(32, 36):
    col_letter = get_column_letter(col)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill

# Pionex
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=36) # Pionex 總資產, AJ列
    cell.fill = bright_orange_fill
    cell.border = Catduck_style

for col in range(37, 41):
    col_letter = get_column_letter(col)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill

# Gate.io
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=41) # Gate.io 總資產, AO列
    cell.fill = lime_green_fill
    cell.border = Catduck_style
for col in range(42, 53):
    col_letter = get_column_letter(col)
    for cell in sheet[col_letter]:
        cell.fill = light_green_fill

# 閒雜人等
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=53) # 閒雜人等1, BA列
    cell.fill = light_purple_fill
    cell.border = Catduck_style
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=54) # 閒雜人等1備註, BB列
    cell.fill = light_green_fill
    cell.border = Catduck_style

for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=55) # 閒雜人等2, BC列
    cell.fill = light_purple_fill
    cell.border = Catduck_style
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=56) # 閒雜人等2備註, BD列
    cell.fill = light_green_fill
    cell.border = Catduck_style


for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=57) # 閒雜人等3, BE列
    cell.fill = light_purple_fill
    cell.border = Catduck_style
for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=58) # 閒雜人等3備註, BF列
    cell.fill = light_green_fill
    cell.border = Catduck_style

for i in range(1, row + 1):
    cell = sheet.cell(row=i, column=59) # 總備註, BG列
    cell.fill = light_orange_fill
    cell.border = Catduck_style





wb.save('crypto_assets.xlsx')

row_number = row
print('--------------------------------------------------------\n')
print(f"此次更新寫入交易所資產表，工作表 {sheet.title} 第 {row_number} 行")
print('--------------------------------------------------------\n')
print('作者：0xCatduck\nhttps://twitter.com/0xCatduck')
print('--------------------------------------------------------\n')
